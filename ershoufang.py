# -*- coding: utf-8 -*-
# author: Frank Zhang
# Email: 565087339@qq.com

from __future__ import print_function
from __future__ import absolute_import
from __future__ import division
from __future__ import unicode_literals
from __future__ import generators
from __future__ import with_statement

import re
from bs4 import BeautifulSoup
from concurrent import futures
import os
import sys
import traceback
import time
import datetime
import pandas as pd
import requests
import json
import shutil
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from fake_useragent import UserAgent
from openpyxl import load_workbook

import smtplib
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.header import Header

############ 全局变量初始化 ##############
HEADERS = dict()
# 并发线程数
NUM_THREADS = None
# 城市选择
city_dict = {
    "成都": "cd",
    "北京": "bj",
    "上海": "sh",
    "广州": "gz",
    "深圳": "sz",
    "南京": "nj",
    "合肥": "hf",
    "杭州": "hz",
}

# 是否打印HTTP错误
PRINT = False
# 伪造User-Agent库初始化
ua = UserAgent()
# 不使用代理
proxies = None
WORKPATH="/home/frank/workspace/lianjia/data"
CITY = city_dict["北京"]


""" HTTP GET 操作封装 """
def get_bs_obj_from_url(http_url):
    done = False
    exception_time = 0
    HEADERS["User-Agent"] = ua.random
    while not done:
        try:
            if PRINT:
                print("正在获取 {}".format(http_url))
            r = requests.get(http_url, headers=HEADERS, proxies=proxies, timeout=3)
            bs_obj = BeautifulSoup(r.text, "lxml")
            done = True
        except Exception as e:
            if PRINT:
                print(e)
            exception_time += 1
            time.sleep(1)
            if exception_time > 10:
                return None
    return bs_obj

""" 判断一个字符串是否可以转成数字 """
def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def esf_mkdir(path):
    path=path.strip()
    path=path.rstrip("\\")
    isExists=os.path.exists(path)

    if not isExists:
        os.makedirs(path)
        print("{} create successfully.".format(path))
        return True
    else:
        print("{} already exist.".format(path))
        return False

def get_district_from_city(city):
    print("---get {} districts---".format(city))
    city_url = "http://{}.lianjia.com".format(city)
    http_url = city_url + "/ershoufang"
    bs_obj = get_bs_obj_from_url(http_url)

    parent_div = bs_obj.find("div", {"data-role": "ershoufang"})
    a_list = parent_div.find_all("a")

    district_list = [a.attrs["href"].replace("/ershoufang/", "")[:-1]
                         for a in a_list
                         if a.attrs['href'].startswith("/ershoufang")]

    print("---total {} districts---".format(len(district_list)))

    return district_list

def get_esf_from_district(city, district):
    http_url = "http://{}.lianjia.com/ershoufang/{}".format(city, district)
    bs_obj = get_bs_obj_from_url(http_url)
    total_esf_num = int(bs_obj.find("h2", {"class": "total fl"}).find("span").get_text())
    print("---district {} total ershoufang numbers: {}---".format(district, total_esf_num))
    esf_list = []

    if total_esf_num == 0:
        print("---district {} total get {}/{}---\n".format(district, len(esf_list), total_esf_num))
        return esf_list

    for price in range(1, 9):
        esf_list_partial = get_esf_id_in_price(city, district, price)
        if esf_list_partial is not None and len(esf_list_partial) > 0:
            esf_list += esf_list_partial
    print("---district {} total get {}/{}---\n".format(district, len(esf_list), total_esf_num))

    return esf_list

def get_esf_id_in_price(city, district, price):
    http_url = "http://{}.lianjia.com/ershoufang/{}/p{}".format(city, district, price)
    bs_obj = get_bs_obj_from_url(http_url)
    total_esf_num = 0
    try:
        total_esf_num = int(bs_obj.find("h2", {"class": "total fl"}).find("span").get_text())
    except Exception as e:
        print("    price {} get error.".format(price))
        pass
    #print("------price {} total : {}---".format(price, total_esf_num))
    esf_list = []

    if total_esf_num == 0:
        print("    price {} finish---done.".format(price))
        return esf_list

    try:
        page_box = bs_obj.find("div", {"class": "page-box house-lst-page-box"})
        total_pages = int(json.loads(page_box.attrs["page-data"])["totalPage"])
    except Exception as e:
        print("    price {} page get error.".format(price))
        return esf_list

    with futures.ThreadPoolExecutor(max_workers=NUM_THREADS) as executor:
        future_list = []
        for page_no in range(1, total_pages + 1):
            future_list.append(executor.submit(get_esf_id_in_page, city, district, price, page_no))
        fail_list = []
        count = 0
        for future in futures.as_completed(future_list):
            page_no, esf_list_partial = future.result()
            if esf_list_partial is None or len(esf_list_partial) == 0:
                fail_list.append(page_no)
            else:
                esf_list += esf_list_partial
                count += 1
                sys.stdout.write("\r    price {} finish {}/{}".format(price, len(esf_list), total_esf_num))
        for page_no in fail_list:
            _, esf_list_partial = get_esf_id_in_page(city, district, price, page_no)
            if esf_list_partial is not None and len(esf_list_partial) > 0:
                esf_list += esf_list_partial
            count += 1
            sys.stdout.write("\r    price {} finish {}/{}".format(price, len(esf_list), total_esf_num))

    print("---done.")
    return esf_list

def get_esf_id_in_page(city, district, price, page_no):
    http_url = "http://{}.lianjia.com/ershoufang/{}/pg{}p{}".format(city, district, page_no, price)
    bs_obj = get_bs_obj_from_url(http_url)

    if bs_obj is None:
        print("get ershoufang id, price {} page {} is none".format(price, page_no))
        return None

    parent_list = bs_obj.find_all("li", {"class": "clear"})

    esf_list = []

    if not (len(parent_list) == 0):
        for li in parent_list:
            esf_url = str(li.find("div", {"class": "title"}).find("a").attrs["href"])
            esf_id = "".join(list(filter(str.isdigit, esf_url)))
            esf_list.append(esf_id)
    return page_no, esf_list


def get_esf_of_city(city):
    district_list = get_district_from_city(city)
    esf_list = []
    for district in district_list:
        esf_of_district = get_esf_from_district(city, district)
        esf_list += esf_of_district
    return esf_list


def get_esf_info(city, esf_id):
    http_url = "https://{}.lianjia.com/ershoufang/{}.html".format(city, esf_id)
    bs_obj = get_bs_obj_from_url(http_url)

    df = pd.DataFrame()

    if bs_obj is not None:
        try:
            test = bs_obj.find("div", {"class": "icon-404 icon fl"})
            if test is not None:
                return esf_id, df

            total_price = bs_obj.find("span", {"class": "total"}).get_text()
            if not is_number(total_price):
                return esf_id, df

            unit_price = bs_obj.find("div", {"class": "unitPrice"}).get_text().replace("元/平米", "")
            huxing = bs_obj.find("div", {"class": "room"}).find("div", {"class": "mainInfo"}).get_text()
            xiaoqu = bs_obj.find("div", {"class": "communityName"}).find("a").get_text()

            area_info = bs_obj.find("div", {"class": "areaName"}).find_all("a")
            chengqu = area_info[0].get_text()
            quyu = area_info[1].get_text()

            base_info = bs_obj.find("div", {"class": "newwrap baseinform"})

            # 基本属性
            base = base_info.find("div", {"class": "base"}).get_text()
            louceng = None if "所在楼层" not in base else base.split("所在楼层")[1].split("(")[0]
            zonglouceng = None if "所在楼层" not in base else base.split("(共")[1].split("层")[0]
            jianzhumianji = None if "建筑面积" not in base else base.split("建筑面积")[1].split("㎡")[0]
            if not is_number(jianzhumianji):
                return esf_id, df

            huxingjiegou = None if "户型结构" not in base else base.split("户型结构")[1].split("\n")[0]
            if "套内面积" not in base:
                taoneimianji = None
            elif "暂无数据" in base.split("套内面积")[1].split("\n")[0]:
                taoneimianji = None
            else:
                taoneimianji = base.split("套内面积")[1].split("㎡")[0]
            jianzhuleixing = None if "建筑类型" not in base else base.split("建筑类型")[1].split("\n")[0]
            chaoxiang = None if "房屋朝向" not in base else base.split("房屋朝向")[1].split("\n")[0]
            jianzhujiegou = None if "建筑结构" not in base else base.split("建筑结构")[1].split("\n")[0]
            zhuangxiu = None if "装修情况" not in base else base.split("装修情况")[1].split("\n")[0]
            tihubili = None if "梯户比例" not in base else base.split("梯户比例")[1].split("\n")[0]
            gongnuan = None if "供暖方式" not in base else base.split("供暖方式")[1].split("\n")[0]
            dianti = None if "配备电梯" not in base else base.split("配备电梯")[1].split("\n")[0]
            chanquan = None if "产权年限" not in base else base.split("产权年限")[1].split("\n")[0]
            yongshui = "商水" if base_info.find(text="商水") is not None else "民水"
            yongdian = "商电" if base_info.find(text="商电") is not None else "民电"

            #　交易属性
            trans = base_info.find("div", {"class": "transaction"}).get_text()
            guapaishijian = None if "挂牌时间" not in trans else trans.split("挂牌时间")[1].split("\n")[0]
            jiaoyiquanshu = None if "交易权属" not in trans else trans.split("交易权属")[1].split("\n")[0]
            fangwuyongtu = None if "房屋用途" not in trans else trans.split("房屋用途")[1].split("\n")[0]
            fangwunianxian = None if "房屋年限" not in trans else trans.split("房屋年限")[1].split("\n")[0]
            chanquansuoshu = None if "产权所属" not in trans else trans.split("产权所属")[1].split("\n")[0]
            diyaxinxi = None if "抵押信息" not in trans else trans.split("抵押信息")[1].split("\n")[0]

            df = pd.DataFrame(index=[esf_id], data=[[http_url, chengqu, quyu, xiaoqu,
                                     huxing, total_price, unit_price, jianzhumianji,
                                     taoneimianji, chaoxiang, louceng, zonglouceng,
                                     huxingjiegou, jianzhuleixing, jianzhujiegou,
                                     fangwuyongtu, jiaoyiquanshu, fangwunianxian,
                                     guapaishijian, zhuangxiu, tihubili, gongnuan,
                                     dianti, chanquan, yongshui, yongdian,
                                     chanquansuoshu, diyaxinxi]],
                              columns=["URL", "城区", "片区", "小区",
                                       "户型", "总价", "单价", "建筑面积",
                                       "套内面积", "朝向", "楼层", "总楼层",
                                       "户型结构", "建筑类型", "建筑结构",
                                       "房屋用途", "交易权属", "房屋年限",
                                       "挂牌时间", "装修", "梯户比例", "供暖",
                                       "配备电梯", "产权", "用水", "用电",
                                       "产权所属", "抵押信息"])
        except Exception as e:
            print("[E]: get_esf_info, esf_id =", esf_id, e)
            traceback.print_exc()
            pass

    return esf_id, df


def get_esf_info_from_esf_list(city, esf_list):
    df_esf_info = pd.DataFrame()
    count = 0
    pct = 0

    with futures.ThreadPoolExecutor(max_workers=NUM_THREADS) as executor:
        future_list = []
        for esf in esf_list:
            future_list.append(executor.submit(get_esf_info, city, esf))
        fail_list = []
        #print(" ")
        for future in futures.as_completed(future_list):
            esf, df_info_partial = future.result()
            if len(df_info_partial) == 0:
                fail_list.append(esf)
            else:
                df_esf_info = df_esf_info.append(df_info_partial)
                count += 1
                sys.stdout.write("\rget ershoufang info: {}/{}".format(count, len(esf_list)))
        for esf in fail_list:
            _, df_info_partial = get_esf_info(city, esf)
            if len(df_info_partial) > 0:
                df_esf_info = df_esf_info.append(df_info_partial)
                count += 1
        sys.stdout.write("\rget ershoufang info: {}/{}".format(count, len(esf_list)))

    print(" ")
    return df_esf_info

def compare_two_list(new_esf_list, old_esf_list):
    add_list = []
    remove_list = []
    same_list = []
    for esf_id in new_esf_list:
        if esf_id not in old_esf_list:
            add_list.append(esf_id)
        else:
            same_list.append(esf_id)
    for esf_id in old_esf_list:
        if esf_id not in new_esf_list:
            remove_list.append(esf_id)
    return add_list, remove_list, same_list

def excel_add_sheet(dataframe, excelwriter, sheetname):
    book = load_workbook(excelwriter.path)
    excelwriter.book = book
    dataframe.to_excel(excelwriter, sheetname, index_label='ID')
    excelwriter.close()
    return

def get_price_changed_esf_info(same_list, new_esf_info, old_esf_info):
    df_jiang = pd.DataFrame()
    df_zhang = pd.DataFrame()
    count = 0
    for esf_id in same_list:
        try:
            new_price = new_esf_info.loc[[esf_id]]["总价"].values[0]
            old_price = old_esf_info.loc[[esf_id]]["总价"].values[0]
            old_unit_price = old_esf_info.loc[esf_id]["单价"]
            new_info = new_esf_info.loc[[esf_id]]
            if new_price > old_price:
                new_info.insert(loc=6, column="原总价", value=old_price)
                new_info.insert(loc=7, column="涨价", value=(new_price-old_price))
                zhangfu=format(((new_price-old_price)/old_price), '.2%')
                new_info.insert(loc=8, column="涨幅", value=zhangfu)
                new_info.insert(loc=10, column="原单价", value=old_unit_price)
                df_zhang = df_zhang.append(new_info)
            elif new_price < old_price:
                new_info.insert(loc=6, column="原总价", value=old_price)
                new_info.insert(loc=7, column="跌价", value=(old_price-new_price))
                diefu=format(((old_price-new_price)/old_price), '.2%')
                new_info.insert(loc=8, column="跌幅", value=diefu)
                new_info.insert(loc=10, column="原单价", value=old_unit_price)
                df_jiang = df_jiang.append(new_info)
            else:
                pass
        except Exception as e:
            print("[E]: get_price_changed, esf_id =", esf_id, e)
            pass
        count += 1
        sys.stdout.write("\rget price change info: {}/{}".format(count, len(same_list)))
    print(" ")
    return df_jiang, df_zhang

def get_chengjiao_yesterday(city):
    district_list = get_district_from_city(city)
    chengjiao = 0
    for district in district_list:
        http_url = 'https://{}.lianjia.com/fangjia/{}'.format(city, district)
        bs_obj = get_bs_obj_from_url(http_url)
        if bs_obj is None:
            chengjiao += 0
            continue
        item = bs_obj.find("div", {"class": "item item-1-2"})
        if item is None:
            chengjiao += 0
            continue
        num = item.find("div", {"class": "num"}).find("span").get_text()
        chengjiao += (0 if "暂无数据" in num else int(num))

    return chengjiao

def get_lianjia_fangjia_info(city):
    http_url = 'https://{}.lianjia.com/fangjia'.format(city)
    bs_obj = get_bs_obj_from_url(http_url)
    if bs_obj is None:
        return 0, 0, 0
    tongji = bs_obj.find("div", {"class": "box-l-b"})
    lj_all = tongji.find_all("div", {"class": "num"})
    lj_new = lj_all[0].get_text()
    lj_ren = lj_all[1].get_text()
    lj_kan = lj_all[2].get_text()

    return lj_new, lj_ren, lj_kan

def get_tongji_info(city, filename):
    lj_new, lj_ren, lj_kan = get_lianjia_fangjia_info(city)
    chengjiao = get_chengjiao_yesterday(city)
    new_str = datetime.date.today().strftime('%Y-%m-%d')

    total_info = pd.read_excel(filename, sheet_name="total", index_col=0)
    total_list = total_info.index.values
    new_info   = pd.read_excel(filename, sheet_name="新上", index_col=0)
    new_list   = new_info.index.values
    rm_info    = pd.read_excel(filename, sheet_name="下架", index_col=0)
    rm_list    = rm_info.index.values
    jiang_info = pd.read_excel(filename, sheet_name="降价", index_col=0)
    jiang_list = jiang_info.index.values
    zhang_info = pd.read_excel(filename, sheet_name="涨价", index_col=0)
    zhang_list = zhang_info.index.values
    junjia     = format(sum(total_info['总价']) * 10000 / sum(total_info['建筑面积']), '.2f')
    jiangfu    = (jiang_info['跌幅'].str.strip("%").astype(float)/100) if len(jiang_list) else 0
    junjiang   = (format(sum(jiangfu) / len(jiangfu), '.2%')) if len(jiang_list) else 0
    zhangfu    = (zhang_info['涨幅'].str.strip("%").astype(float)/100) if len(zhang_list) else 0
    junzhang   = (format(sum(zhangfu) / len(zhangfu), '.2%')) if len(zhang_list) else 0

    info = pd.DataFrame(index=[new_str],
                        data=[[len(total_list), junjia, chengjiao,
                               len(new_list), len(rm_list), len(jiang_list),
                               junjiang, len(zhang_list), junzhang, lj_new,
                               lj_ren, lj_kan]],
                        columns=['总数', '均价', '成交', '上架', '下架',
                                 '降价', '降幅', '涨价', '涨幅', '新上',
                                 '新客户', '带看'])

    return info

def get_email_content(info):
    content = '本期统计信息：\n'
    content += '线上总套数：{}套，'.format(info['总数'].values[0])
    content += '均价：{}元/平米\n'.format(info['均价'].values[0])
    content += '昨日成交数：{}套\n'.format(info['成交'].values[0])
    content += '新上房源数：{}套\n'.format(info['上架'].values[0])
    content += '下架房源数：{}套\n'.format(info['下架'].values[0])
    content += '降价房源数：{}套，'.format(info['降价'].values[0])
    content += '均降：{}\n'.format(info['降幅'].values[0])
    content += '涨价房源数：{}套，'.format(info['涨价'].values[0])
    content += '均涨：{}\n'.format(info['涨幅'].values[0])
    content += '\n'
    content += '链家统计信息：\n'
    content += '新增房源数：{}套\n'.format(info['新上'].values[0])
    content += '新增客户数：{}人\n'.format(info['新客户'].values[0])
    content += '新增带看数：{}次\n'.format(info['带看'].values[0])

    return content

def addimg(src, imgid):
    fp = open(src, 'rb')
    msgImage = MIMEImage(fp.read())
    fp.close()
    msgImage.add_header('Content-ID', imgid)
    return msgImage

def send_email(content, filename):
    sender = '565087339@qq.com'
    receivers = ['565087339@qq.com']
    key = open('../key', 'r').read()

    message = MIMEMultipart()
    message['From'] = sender
    message['Subject'] = Header(filename, 'utf-8')
    #message.attach(MIMEText(content, 'plain', 'utf-8'))
    html = '<p>{}</p><p><img src="cid:image1"></p>'.format(content.replace('\n', '<br>'))
    message.attach(MIMEText(html, 'html', 'utf-8'))
    message.attach(addimg("total.jpg","image1"))

    att = MIMEText(open(filename, 'rb').read(), 'base64', 'utf-8')
    att["Content-Type"] = 'application/octet-stream'
    att_str = 'attachment; filename={}'.format(filename)
    att["Content-Disposition"] = att_str
    message.attach(att)

    try:
        smtpObj = smtplib.SMTP('smtp.qq.com')
        smtpObj.login(sender, key)
        smtpObj.sendmail(sender, receivers, message.as_string())
        print("send email successfully.")
    except smtplib.SMTPException:
        print("send email failed.")
    return

def get_tongji_plot(filename):
    info = pd.read_excel(filename, sheet_name="统计", index_col=0)
    info = info.sort_index()
    try:
        info.plot(x=pd.to_datetime(info.index), y=['总数', '均价', '成交'],
                  marker='o', subplots=True, grid=True, figsize=(12,6))
        #plt.title('北京可售房源总数')
        plt.savefig('total.jpg')
    except Exception as e:
        print("get tongji plot failed", e)
    return

def get_esf_location_by_index(index, http_url):
    #print("index {} start".format(index))
    lng = 0.0
    lat = 0.0
    bs_obj = get_bs_obj_from_url(http_url)
    if bs_obj is None:
        print("get location failed, index={}".format(index))
        return index, lng, lat
    try:
        lng = float(bs_obj.find('lng').get_text())
        lat = float(bs_obj.find('lat').get_text())
    except Exception as e:
        print("get lng/lat failed. bs_obj={}".format(bs_obj))
        pass
    #print("index {} end".format(index))

    return index, lng, lat

def get_esf_location(filename):
    ak = open('../ak', 'r').read().replace('\n', '')
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name('total')
    max_row = ws.max_row
    max_col = ws.max_column
    ws.cell(row=1, column=max_col+1, value='经度')
    ws.cell(row=1, column=max_col+2, value='纬度')
    count = 0

    with futures.ThreadPoolExecutor(max_workers=None) as executor:
        future_list = []
        for index in range(2, max_row+1):
            chengqu = ws.cell(row=index, column=3).value
            xiaoqu  = ws.cell(row=index, column=5).value
            location = '北京市{}区{}'.format(chengqu, xiaoqu)
            if location is None:
                print("get location failed, index={}".format(index))
                continue
            http_url = 'http://api.map.baidu.com/geocoder/v2/?address={}&ak={}'.format(location, ak)
            future_list.append(executor.submit(get_esf_location_by_index, index, http_url))
        fail_list = []
        for future in futures.as_completed(future_list):
            idx, lng, lat = future.result()
            if lng == 0.0:
                fail_list.append(idx)
            else:
                ws.cell(row=idx, column=max_col+1).value=format(lng, '.6f')
                ws.cell(row=idx, column=max_col+2).value=format(lat, '.6f')
                count += 1
                sys.stdout.write("\rget location info: {}/{}...".format(count, max_row-1))
        for idx in fail_list:
            chengqu = ws.cell(row=index, column=3).value
            xiaoqu  = ws.cell(row=index, column=5).value
            location = '北京市{}区{}'.format(chengqu, xiaoqu)
            if location is None:
                print("get location failed, index={}".format(index))
                continue
            http_url = 'http://api.map.baidu.com/geocoder/v2/?address={}&ak={}'.format(location, ak)
            _, lng, lat = get_esf_location_by_index(idx, http_url)
            ws.cell(row=idx, column=max_col+1).value=format(lng, '.6f')
            ws.cell(row=idx, column=max_col+2).value=format(lat, '.6f')
            count += 1
            sys.stdout.write("\rget location info: {}/{}...".format(count, max_row-1))
        print("done.")

    wb.save(filename)
    return



def main():
    ###########################################################
    # 总共N个步骤，依次运行。
    # 运行第一步的时候，把其余几步的代码注释掉，依次类推
    ###########################################################

    os.chdir(WORKPATH)
    if not PRINT:
        log_file = open('../log', 'a')
        sys.stdout = log_file

    # 1. make new dir
    print("\n1. getting date info...")
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    new_str = today.strftime('%Y-%m-%d')
    old_str = yesterday.strftime('%Y-%m-%d')
    new_file = "{}_info_{}.xlsx".format(CITY, new_str)
    old_file = "{}_info_{}.xlsx".format(CITY, old_str)
    print("today: {}, yesterday: {}.".format(new_str, old_str))

    # 2. get ershoufang id of the city
    print("\n2.getting ershoufang list...")
    esf_list = get_esf_of_city(CITY)
    with open("{}_list_{}.txt".format(CITY, new_str), mode="w") as f:
        for esf in esf_list:
            f.write(esf + "\n")
    print("ershoufang list write finished.")

    # 3. get ershoufang info
    print("\n3. getting ershoufang info...")
    with open("{}_list_{}.txt".format(CITY, new_str), mode="r") as f:
        esf_list = [int(line[:-1]) for line in f.readlines()]
    print("get ershoufang info start...")
    df_esf_info = get_esf_info_from_esf_list(CITY, esf_list)
    writer = pd.ExcelWriter(new_file)
    df_esf_info.to_excel(writer, "total")
    writer.save()
    try:
       os.remove("{}_list_{}.txt".format(CITY, new_str))
    except Exception as e:
        pass
    print("ershoufang info write finished.")

    # 4. find new ershoufang list and info
    print("\n4. getting different ershoufang list...")
    df_esf_info = pd.read_excel(new_file, sheet_name="total", index_col=0)
    new_esf_list = df_esf_info.index.values
    df_esf_info = pd.read_excel(old_file, sheet_name="total", index_col=0)
    old_esf_list = df_esf_info.index.values
    add_list, remove_list, same_list = compare_two_list(new_esf_list, old_esf_list)
    print("different ershoufang list finished.")

    # 5. get new ershoufang today
    print("\n5. getting new ershoufang info...")
    df_esf_info = pd.read_excel(new_file, sheet_name="total", index_col=0)
    df_esf_added = df_esf_info.loc[add_list]
    writer = pd.ExcelWriter(new_file)
    excel_add_sheet(df_esf_added, writer, "新上")
    print("new ershoufang info write finished.")

    # 6. get removed ershoufang today
    print("\n6. getting removed ershoufang info...")
    df_esf_info = pd.read_excel(old_file, sheet_name="total", index_col=0)
    df_esf_removed = df_esf_info.loc[remove_list]
    writer = pd.ExcelWriter(new_file)
    excel_add_sheet(df_esf_removed, writer, "下架")
    print("removed ershoufang info write finished.")

    # 7. get price changed ershoufang today
    print("\n7. getting price changed ershoufang info...")
    new_esf_info = pd.read_excel(new_file, sheet_name="total", index_col=0)
    old_esf_info = pd.read_excel(old_file, sheet_name="total", index_col=0)
    df_jiang, df_zhang = get_price_changed_esf_info(same_list, new_esf_info, old_esf_info)
    writer = pd.ExcelWriter(new_file)
    excel_add_sheet(df_jiang, writer, "降价")
    excel_add_sheet(df_zhang, writer, "涨价")
    print("price changed ershoufang info write finished.")

    # 8. get statistical information
    print("\n8. getting statistical information")
    info = get_tongji_info(CITY, new_file)
    old_info = pd.read_excel(old_file, sheet_name="统计", index_col=0)
    info = info.append(old_info)
    writer = pd.ExcelWriter(new_file)
    writer.book = load_workbook(writer.path)
    info.to_excel(writer, "统计", index_label='日期')
    writer.close()
    print("statistical information finished.")

    # 9. get plot of statistical information
    print("\n9. getting plot of statistical information")
    get_tongji_plot(new_file)
    print("statistical information plot finished.")

    # 10. send email with the new file
    print("\n10. sending email with the data...")
    #info = get_tongji_info(CITY, new_file)
    info = pd.read_excel(new_file, sheet_name='统计', index_col=0)
    content = get_email_content(info)
    send_email(content, new_file)
    print("send email finished.")

    '''
    # 11. getting location information
    print("\n11. getting location information...")
    get_esf_location(new_file)
    print("get location finished.")
    '''

if __name__ == "__main__":
    main()
